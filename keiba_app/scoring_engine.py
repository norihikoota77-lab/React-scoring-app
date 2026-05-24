import os
import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)


class ScoringEngine:
    """
    Excel ファイルの読み込み・正誤判定・
    Excel レポート出力を担うクラス。
    """

    # ── ランク閾値 ──────────────────────────────

    RANK_THRESHOLDS = {
        "S": 100,
        "A": 70,
        "B": 50,
        "C": 0,
    }

    RESULT_MESSAGES = {
        "S": ("🌟🏆 [G1制覇] 伝説の三冠馬級！", "#FFD700"),
        "A": ("🥈 [重賞入着] 素晴らしい末脚です", "#C0C0C0"),
        "B": ("🐎 [入賞] 掲示板に載りました", "#FFCC99"),
        "C": ("🏃 [未勝利] ゲート練習からやり直し", "#A9A9A9"),
    }

    def __init__(self):

        self.correct_map: dict = {}
        self.user_map: dict = {}

        self.score: int = 0
        self.valid_count: int = 0
        self.percentage: float = 0.0

        self.rows_data: list = []
        self.judgments: dict = {}

        # 新規追加
        self.exam_name: str = ""
        self.user_name: str = ""

    # ── 試験情報取得 ────────────────────────────

    @staticmethod
    def get_exam_info(
        file_path: str
    ) -> tuple[str, str]:

        """
        A13 → 試験名
        A14 → 受験者名
        """

        try:

            df = pd.read_excel(
                file_path,
                header=None
            )

            exam_name = str(
                df.iloc[12, 0]
            ).strip()

            user_name = str(
                df.iloc[13, 0]
            ).strip()

            return exam_name, user_name

        except Exception as e:

            raise RuntimeError(
                f"試験情報取得エラー: {e}"
            ) from e

    # ── 試験整合性チェック ───────────────────────

    def validate_exam(
        self,
        correct_file: str,
        user_file: str
    ) -> tuple[str, str]:

        correct_exam, _ = self.get_exam_info(
            correct_file
        )

        user_exam, user_name = self.get_exam_info(
            user_file
        )

        if correct_exam != user_exam:

            raise ValueError(
                "異なる試験データです\n"
                f"正解マスタ: {correct_exam}\n"
                f"解答データ: {user_exam}"
            )

        return user_exam, user_name

    # ── Excel 読み込み ─────────────────────────

    @staticmethod
    def load_answers(file_path: str) -> dict:
        """
        Excel ファイルを読み込んで
        {問番号: 答え文字列} の辞書を返す。
        """

        try:

            df = pd.read_excel(
                file_path,
                header=None
            )

            data_map: dict = {}

            col_pairs = [
                (0, 1),
                (2, 3),
                (4, 5),
                (6, 7)
            ]

            # 1行目が数値でなければ
            # ヘッダー行とみなしてスキップ

            start_row = (
                1
                if not str(df.iloc[0, 0]).isdigit()
                else 0
            )

            for col_num_idx, col_ans_idx in col_pairs:

                for row_idx in range(10):

                    try:

                        target_row = (
                            start_row + row_idx
                        )

                        if target_row >= len(df):
                            continue

                        q_val = df.iloc[
                            target_row,
                            col_num_idx
                        ]

                        ans_val = df.iloc[
                            target_row,
                            col_ans_idx
                        ]

                        if pd.notna(q_val):

                            q_num = int(float(q_val))

                            s_val = (
                                str(ans_val).split(".")[0]
                                if isinstance(ans_val, float)
                                else str(ans_val)
                            )

                            data_map[q_num] = (
                                s_val.strip().upper()
                                if pd.notna(ans_val)
                                and s_val != "nan"
                                else None
                            )

                    except Exception:
                        continue

            return data_map

        except Exception as e:

            raise RuntimeError(
                f"エクセル読込エラー: {e}"
            ) from e

    # ── 採点メイン ─────────────────────────────

    def grade(
        self,
        correct_file: str,
        user_file: str
    ) -> None:

        # 試験整合性チェック

        self.exam_name, self.user_name = (
            self.validate_exam(
                correct_file,
                user_file
            )
        )

        self.correct_map = self.load_answers(
            correct_file
        )

        self.user_map = self.load_answers(
            user_file
        )

        all_qs = sorted(
            self.correct_map.keys()
        )

        self.rows_data = []
        self.judgments = {}

        self.score = 0
        self.valid_count = 0

        for q in all_qs:

            c_ans = self.correct_map[q]

            u_ans = self.user_map.get(q)

            is_valid = c_ans is not None

            is_correct = (
                str(u_ans) == str(c_ans)
            ) if is_valid else False

            if is_valid:

                self.valid_count += 1

                if is_correct:
                    self.score += 1

            self.judgments[q] = (
                is_correct,
                is_valid
            )

            self.rows_data.append([
                q,
                u_ans if u_ans else "未記入",
                c_ans if is_valid else "-",
                (
                    "⭕"
                    if (is_valid and is_correct)
                    else (
                        "✖"
                        if is_valid
                        else "-"
                    )
                ),
            ])

        self.percentage = (
            (
                self.score
                / self.valid_count
                * 100
            )
            if self.valid_count > 0
            else 0.0
        )

    # ── ランク判定 ─────────────────────────────

    def get_rank(self) -> str:

        if self.percentage == 100:
            return "S"

        elif self.percentage >= 70:
            return "A"

        elif self.percentage >= 50:
            return "B"

        else:
            return "C"

    def get_result_message(self):

        return self.RESULT_MESSAGES[
            self.get_rank()
        ]

    # ── Excel レポート出力 ─────────────────────

    def export_excel(self, output_file: str) -> None:

        left_df = pd.DataFrame(
            self.rows_data[:20],
            columns=["問題", "解答", "正解", "判定"]
        )

        right_df = pd.DataFrame(
            self.rows_data[20:40],
            columns=["問題", "解答", "正解", "判定"]
        )

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            # 左テーブル
            left_df.to_excel(
                writer,
                sheet_name="レース結果",
                startrow=5,
                startcol=0,
                index=False
            )

            # 右テーブル
            right_df.to_excel(
                writer,
                sheet_name="レース結果",
                startrow=5,
                startcol=5,
                index=False
            )

            ws = writer.sheets["レース結果"]

            # ======================
            # タイトル情報
            # ======================

            ws["A1"] = "🐎 競馬スコアラー"

            ws["A2"] = f"試験名：{self.exam_name}"
            ws["A3"] = f"受験者：{self.user_name}"

            ws["G2"] = f"得点率：{round(self.percentage, 1)}%"
            ws["G3"] = f"ランク：{self.get_rank()}"

            # ======================
            # 列幅
            # ======================

            widths = {
                "A": 8,
                "B": 10,
                "C": 10,
                "D": 8,

                "F": 8,
                "G": 10,
                "H": 10,
                "I": 8,
            }

            for col, width in widths.items():
                ws.column_dimensions[col].width = width


            self._apply_excel_styles(ws)
 
    # ── Excelスタイル適用 ──────────────────────
    def _apply_excel_styles(self, ws) -> None:

        all_qs = sorted(self.correct_map.keys())

        header_fill = PatternFill(
            start_color="DA1F28",
            end_color="DA1F28",
            fill_type="solid"
        )

        ok_fill = PatternFill(
            start_color="E6FFFA",
            end_color="E6FFFA",
            fill_type="solid"
        )

        ng_fill = PatternFill(
            start_color="FFEBEE",
            end_color="FFEBEE",
            fill_type="solid"
        )

        answer_fill = PatternFill(
            start_color="FFF8DC",
            end_color="FFF8DC",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        START_ROW = 6

        target_cols = [1, 2, 3, 4, 6, 7, 8, 9]

        for r in range(START_ROW, START_ROW + 21):

            for c in target_cols:

                cell = ws.cell(row=r, column=c)

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                # ヘッダー
                if r == START_ROW:

                    cell.fill = header_fill

                    cell.font = Font(
                        color="FFFFFF",
                        bold=True
                    )

                else:

                    q_idx = (
                        0 if c <= 4 else 20
                    ) + (r - (START_ROW + 1))

                    if q_idx < len(all_qs):

                        is_ok, is_valid = self.judgments[
                            all_qs[q_idx]
                        ]

                        # 解答列
                        if c in (2, 7):

                            cell.fill = answer_fill

                        # 正誤色
                        elif is_valid:

                            cell.fill = (
                                ok_fill
                                if is_ok
                                else ng_fill
                            )

                        # 判定列
                        if c in (4, 9):

                            cell.font = Font(
                                color=(
                                    "0000FF"
                                    if is_ok
                                    else "FF0000"
                                ),
                                bold=True,
                                size=14
                            )